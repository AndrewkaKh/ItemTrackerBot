import os
import pandas as pd
from openpyxl import load_workbook
from config import REPORT_FOLDER

def change_width_column(sheet):
    for col in sheet.iter_cols(1, sheet.max_column):
        if col[0].value == "Наименование":
            for cell in col:
                cell.alignment = cell.alignment.copy(wrap_text=True)
            sheet.column_dimensions[col[0].column_letter].width = 60
        else:
            for cell in col:
                cell.alignment = cell.alignment.copy(wrap_text=True)
            sheet.column_dimensions[col[0].column_letter].width = 20


def generate_excel(stock_data, movement_data, stock_pay_user_data, semi_finished_products, products):
    """
    Генерация Excel отчета.
    :param stock_data: Список остатков на складе (формат: [{'field': value}, ...])
    :param movement_data: Список движения товаров (формат: [{'field': value}, ...])
    :param stock_pay_user_data: Список остатка денежных средств сотрудников (формат: [{'field': value}, ...])
    :param semi_finished_products: Список полуфабрикатов (формат: [{'field': value}, ...])
    :param products: Список товаров (формат: [{'field': value}, ...])
    """
    try:
        os.makedirs(REPORT_FOLDER, exist_ok=True)

        stock_df = pd.DataFrame(stock_data)
        movement_df = pd.DataFrame(movement_data)
        stock_pay_user_df = pd.DataFrame(stock_pay_user_data)
        semi_finished_products_df = pd.DataFrame(semi_finished_products)
        products_df = pd.DataFrame(products)

        if "Артикул" in stock_df.columns:
            stock_df = stock_df.sort_values(by="Артикул")

        report_file = os.path.join(REPORT_FOLDER, "warehouse_report.xlsx")
        with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
            stock_df.to_excel(writer, sheet_name="Stock", index=False)
            movement_df.to_excel(writer, sheet_name="Movements", index=False)
            stock_pay_user_df.to_excel(writer, sheet_name="Accounts Balances", index=False)
            semi_finished_products_df.to_excel(writer, sheet_name="Semi-Finished Products", index=False)
            products_df.to_excel(writer, sheet_name="Products", index=False)

        sheets = ["Stock", "Movements", "Accounts Balances", "Semi-Finished Products", "Products"]

        wb = load_workbook(report_file)

        for name_sheet in sheets:
            sheet = wb[name_sheet]
            change_width_column(sheet)

        wb.save(report_file)

        return report_file
    except Exception as e:
        raise ValueError(f"Ошибка при генерации Excel отчета: {str(e)}")


def generate_excel_for_movement(movement_data):
    """
    Генерация Excel отчета.
    :param movement_data: Список движения товаров за какой-то отрезок времени(формат: [{'field': value}, ...])
    """
    try:
        os.makedirs(REPORT_FOLDER, exist_ok=True)

        report_file = os.path.join(REPORT_FOLDER, "warehouse_report.xlsx")
        with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
            movement_data.to_excel(writer, sheet_name="Movements", index=False)
        return report_file
    except Exception as e:
        raise ValueError(f"Ошибка при генерации Excel отчета(filter): {str(e)}")